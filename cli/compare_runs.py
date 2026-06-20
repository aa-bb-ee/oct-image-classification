from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

DEFAULT_OUTPUT_ROOT = Path("experiment_outputs")
DEFAULT_OUTPUT_CSV = Path("reports/model_comparison.csv")
DEFAULT_OUTPUT_MD = Path("reports/model_comparison.md")
DEFAULT_OUTPUT_XLSX = Path("reports/model_comparison.xlsx")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Aggregate OCT experiment summary JSON files into comparison reports.",
    )
    parser.add_argument("--output_root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--output_csv", type=Path, default=DEFAULT_OUTPUT_CSV)
    parser.add_argument("--output_md", type=Path, default=DEFAULT_OUTPUT_MD)
    parser.add_argument("--output_xlsx", type=Path, default=DEFAULT_OUTPUT_XLSX)
    parser.add_argument(
        "--sort_metric",
        default="test_accuracy",
        choices=[
            "test_accuracy",
            "test_macro_precision",
            "test_macro_recall",
            "test_macro_f1",
            "test_weighted_precision",
            "test_weighted_recall",
            "test_weighted_f1",
            "test_roc_auc_macro",
            "test_roc_auc_weighted",
            "val_macro_precision",
            "val_macro_recall",
            "val_macro_f1",
            "val_weighted_precision",
            "val_weighted_recall",
            "val_weighted_f1",
            "val_roc_auc_macro",
            "val_roc_auc_weighted",
        ],
    )
    return parser.parse_args()


def _get_float(payload: dict[str, Any], *path: str) -> float | None:
    cur: Any = payload
    for key in path:
        if not isinstance(cur, dict) or key not in cur:
            return None
        cur = cur[key]
    if isinstance(cur, (int, float)):
        return float(cur)
    return None


def _load_json(path: Path) -> dict[str, Any]:
    """
    Lädt eine JSON-Datei robust.

    Wird verwendet für:
    - *_summary.json
    - run_config.json

    Falls die Datei nicht existiert, wird ein leeres Dict zurückgegeben.
    """
    if not path.exists():
        return {}

    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _load_run_config(summary_path: Path) -> dict[str, Any]:
    return _load_json(summary_path.parent / "run_config.json")


def _load_classification_report_csv(path: Path) -> dict[str, dict[str, Any]]:
    """
    NEU:
    Lädt den von reporting.py exportierten Classification Report.

    reporting.py schreibt:
    - <run_id>_test_report.csv
    - <run_id>_validation_report.csv

    Diese Dateien enthalten Zeilen wie:
    CNV, DME, DRUSEN, NORMAL, accuracy, macro avg, weighted avg

    Für den Modellvergleich brauchen wir vor allem die echten Klassenzeilen.
    """
    if not path.exists():
        return {}

    rows: dict[str, dict[str, Any]] = {}

    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)

        for row in reader:
            # Pandas speichert den Index beim to_csv(index=True) meist
            # in einer leeren ersten Spalte mit Namen "".
            label = (
                row.get("")
                or row.get("Unnamed: 0")
                or row.get("index")
                or row.get("label")
            )

            if not label:
                continue

            rows[label] = row

    return rows


def _csv_float(
    report: dict[str, dict[str, Any]],
    label: str,
    metric: str,
) -> float | None:
    """
    NEU:
    Liest einen Float aus dem Classification-Report-CSV.

    Beispiel:
    _csv_float(test_report, "CNV", "precision")
    """
    if label not in report:
        return None

    value = report[label].get(metric)

    if value in (None, ""):
        return None

    try:
        return float(value)
    except ValueError:
        return None


def _safe_column_name(value: str) -> str:
    """
    NEU:
    Macht Klassennamen robust für CSV-Spaltennamen.

    Beispiel:
    'macro avg' -> 'macro_avg'
    """
    return (
        value.strip()
        .replace(" ", "_")
        .replace("-", "_")
        .replace("/", "_")
        .lower()
    )


def _with_sort_metric_suffix(path: Path, sort_metric: str) -> Path:
    """
    Ergänzt die Sortiermetrik im Dateinamen.

    Beispiel:
    reports/model_comparison.csv
    -> reports/model_comparison_sorted_by_test_accuracy.csv
    """
    return path.with_name(
        f"{path.stem}_sorted_by_{sort_metric}{path.suffix}"
    )


def _add_per_class_metrics(
    row: dict[str, Any],
    report: dict[str, dict[str, Any]],
    class_names: list[str],
    prefix: str,
) -> None:
    """
    NEU:
    Ergänzt Precision/Recall/F1/Support pro Klasse.

    Beispielspalten für prefix='test':
    - test_CNV_precision
    - test_CNV_recall
    - test_CNV_f1
    - test_CNV_support

    Beispielspalten für prefix='val':
    - val_CNV_precision
    - val_CNV_recall
    - val_CNV_f1
    - val_CNV_support
    """
    for class_name in class_names:
        col_class = _safe_column_name(class_name)

        row[f"{prefix}_{col_class}_precision"] = _csv_float(
            report,
            class_name,
            "precision",
        )
        row[f"{prefix}_{col_class}_recall"] = _csv_float(
            report,
            class_name,
            "recall",
        )
        row[f"{prefix}_{col_class}_f1"] = _csv_float(
            report,
            class_name,
            "f1-score",
        )
        row[f"{prefix}_{col_class}_support"] = _csv_float(
            report,
            class_name,
            "support",
        )


def _get_run_dir(summary_path: Path) -> Path:
    """
    Bestimmt den Run-Ordner aus dem Pfad zur Summary.
    """
    return summary_path.parents[2]


def _find_report_path(
    summary_path: Path,
    run_id: str,
    kind: str,
) -> Path:
    """
    Erzeugt den Pfad zum Test- oder Validation-Report.
    """
    return summary_path.parent / f"{run_id}_{kind}_report.csv"


def load_rows(output_root: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for summary_path in sorted(output_root.rglob("*_summary.json")):
        summary = _load_json(summary_path)
        run_config = _load_run_config(summary_path)

        run_dir = _get_run_dir(summary_path)
        run_id = summary.get("run_id", run_dir.name)
        model_path = run_dir / "models" / "best_model.keras"

        class_names = summary.get("class_names", [])
        if not isinstance(class_names, list):
            class_names = []

        test_report_path = _find_report_path(
            summary_path=summary_path,
            run_id=run_id,
            kind="test",
        )
        validation_report_path = _find_report_path(
            summary_path=summary_path,
            run_id=run_id,
            kind="validation",
        )

        # NEU:
        # Per-class Reports aus CSV laden.
        test_report = _load_classification_report_csv(test_report_path)
        validation_report = _load_classification_report_csv(validation_report_path)

        row: dict[str, Any] = {
            # ============================================================
            # 1) Run metadata
            # ============================================================
            "run_id": run_id,
            "run_name": summary.get("run_name", run_config.get("run_name")),
            "model_name": summary.get("model_name", run_config.get("model_name")),

            # ============================================================
            # 2) Training setup / Hyperparameter
            # ============================================================
            "train_mode": run_config.get("train_mode"),
            "base_model_path": run_config.get("base_model_path"),
            "fine_tune": run_config.get("fine_tune"),
            "unfreeze_last_n": run_config.get("unfreeze_last_n"),
            "use_class_weights": run_config.get("use_class_weights"),
            "use_augmentation": run_config.get("use_augmentation"),
            "dropout": run_config.get("dropout"),
            "learning_rate": run_config.get("learning_rate"),
            "fine_tune_lr": run_config.get("fine_tune_lr"),
            "batch_size": run_config.get("batch_size"),
            "epochs": run_config.get("epochs"),
            "fine_tune_epochs": run_config.get("fine_tune_epochs"),
            "img_size": run_config.get("img_size"),
            "seed": run_config.get("seed"),
            "val_split": run_config.get("val_split"),

            # ============================================================
            # 3) Aggregated TEST metrics
            # ============================================================
            "test_accuracy": _get_float(
                summary,
                "test_results",
                "manual_test_accuracy",
            ),
            "test_macro_precision": _get_float(
                summary,
                "test_results",
                "macro_precision",
            ),
            "test_macro_recall": _get_float(
                summary,
                "test_results",
                "macro_recall",
            ),
            "test_macro_f1": _get_float(
                summary,
                "test_results",
                "macro_f1",
            ),
            "test_weighted_precision": _get_float(
                summary,
                "test_results",
                "weighted_precision",
            ),
            "test_weighted_recall": _get_float(
                summary,
                "test_results",
                "weighted_recall",
            ),
            "test_weighted_f1": _get_float(
                summary,
                "test_results",
                "weighted_f1",
            ),
            "test_roc_auc_macro": _get_float(
                summary,
                "test_results",
                "roc_auc_ovr_macro",
            ),
            "test_roc_auc_weighted": _get_float(
                summary,
                "test_results",
                "roc_auc_ovr_weighted",
            ),
            "test_mean_entropy": _get_float(
                summary,
                "test_results",
                "test_mean_entropy",
            ),
            "test_mean_normalized_entropy": _get_float(
                summary,
                "test_results",
                "test_mean_normalized_entropy",
            ),

            # ============================================================
            # 4) Aggregated VALIDATION metrics
            # ============================================================
            "val_macro_precision": _get_float(
                summary,
                "validation_results",
                "val_macro_precision",
            ),
            "val_macro_recall": _get_float(
                summary,
                "validation_results",
                "val_macro_recall",
            ),
            "val_macro_f1": _get_float(
                summary,
                "validation_results",
                "val_macro_f1",
            ),
            "val_weighted_precision": _get_float(
                summary,
                "validation_results",
                "val_weighted_precision",
            ),
            "val_weighted_recall": _get_float(
                summary,
                "validation_results",
                "val_weighted_recall",
            ),
            "val_weighted_f1": _get_float(
                summary,
                "validation_results",
                "val_weighted_f1",
            ),
            "val_roc_auc_macro": _get_float(
                summary,
                "validation_results",
                "val_roc_auc_ovr_macro",
            ),
            "val_roc_auc_weighted": _get_float(
                summary,
                "validation_results",
                "val_roc_auc_ovr_weighted",
            ),
            "val_mean_entropy": _get_float(
                summary,
                "validation_results",
                "val_mean_entropy",
            ),
            "val_mean_normalized_entropy": _get_float(
                summary,
                "validation_results",
                "val_mean_normalized_entropy",
            ),
        }

        # ================================================================
        # 5) Per-class TEST metrics
        # ================================================================
        _add_per_class_metrics(
            row=row,
            report=test_report,
            class_names=class_names,
            prefix="test",
        )

        # ================================================================
        # 6) Per-class VALIDATION metrics
        # ================================================================
        _add_per_class_metrics(
            row=row,
            report=validation_report,
            class_names=class_names,
            prefix="val",
        )

        # ================================================================
        # 7) Paths and existence checks
        # ================================================================
        row.update(
            {
                "summary_path": str(summary_path),
                "test_report_path": str(test_report_path),
                "validation_report_path": str(validation_report_path),
                "model_path": str(model_path),
                "model_exists": model_path.exists(),

                # Internes Feld, wird nicht in CSV geschrieben.
                "_sort_metric": row.get("test_accuracy"),
            }
        )

        rows.append(row)

    return rows


def sort_rows(rows: list[dict[str, Any]], metric: str) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            row.get(metric) is None,
            -(row.get(metric) or float("-inf")),
            str(row.get("run_id", "")),
        ),
    )


def write_csv(rows: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return

    fieldnames = [key for key in rows[0].keys() if not key.startswith("_")]
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key) for key in fieldnames})


def write_xlsx(rows: list[dict[str, Any]], path: Path) -> None:
    """
    Schreibt den Modellvergleich als echte Excel-Datei.

    Vorteil gegenüber CSV:
    - Zahlen bleiben echte Zahlen
    - keine Probleme mit Dezimalpunkt/Dezimalkomma
    - Filter, Freeze Header und Farbskalen werden direkt gesetzt
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Model Comparison"

    if not rows:
        wb.save(path)
        return

    fieldnames = [key for key in rows[0].keys() if not key.startswith("_")]

    ws.append(fieldnames)

    for row in rows:
        ws.append([row.get(key) for key in fieldnames])

    # Header formatieren
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            text_rotation=45,
            vertical="bottom",
            horizontal="center",
        )

    # Header einfrieren und Filter aktivieren
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Spaltenbreiten + Zahlenformat
    metric_columns: list[str] = []

    for col_idx, column_name in enumerate(fieldnames, start=1):
        col_letter = get_column_letter(col_idx)

        if column_name.endswith("_path"):
            ws.column_dimensions[col_letter].width = 45
        elif column_name in {"run_id", "run_name"}:
            ws.column_dimensions[col_letter].width = 38
        elif column_name == "model_name":
            ws.column_dimensions[col_letter].width = 18
        else:
            ws.column_dimensions[col_letter].width = 16

        # Metrikspalten erkennen
        is_metric_col = (
            column_name.startswith("test_")
            or column_name.startswith("val_")
        ) and not column_name.endswith("_path")

        is_support_col = column_name.endswith("_support")

        if is_metric_col:
            metric_columns.append(col_letter)

            for row_idx in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row_idx}"]

                if isinstance(cell.value, float):
                    if is_support_col:
                        cell.number_format = "0"
                    else:
                        cell.number_format = "0.0000"

    # Farbskalen für Metriken:
    # Bei Accuracy/F1/Recall/Precision/ROC-AUC ist höher besser.
    for col_letter in metric_columns:
        header = ws[f"{col_letter}1"].value

        if header is None:
            continue

        header = str(header)

        # Entropie: niedriger ist besser -> Grün niedrig, Rot hoch.
        if "entropy" in header:
            rule = ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            )
        else:
            rule = ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            )

        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{ws.max_row}",
            rule,
        )

    wb.save(path)


def _fmt(value: Any) -> str:
    if isinstance(value, float):
        return f"{value:.4f}"
    if value is None:
        return "-"
    return str(value)


def write_markdown(rows: list[dict[str, Any]], path: Path) -> None:
    """
    Schreibt eine kompakte Markdown-Übersicht.

    Hinweis:
    Die Markdown-Datei ist bewusst kürzer als die CSV.
    Die vollständigen per-class Metriken stehen in der CSV.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "rank",
        "run_id",
        "model",
        "mode",
        "accuracy",
        "macro_precision",
        "macro_recall",
        "macro_f1",
        "weighted_f1",
        "roc_auc",
        "val_f1",
        "class_weights",
        "augmentation",
    ]
    lines = [
        "# Model Comparison",
        "",
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]

    for idx, row in enumerate(rows, start=1):
        lines.append(
            "| "
            + " | ".join(
                [
                    str(idx),
                    str(row.get("run_id")),
                    str(row.get("model_name")),
                    _fmt(row.get("train_mode")),
                    _fmt(row.get("test_accuracy")),
                    _fmt(row.get("test_macro_precision")),
                    _fmt(row.get("test_macro_recall")),
                    _fmt(row.get("test_macro_f1")),
                    _fmt(row.get("test_weighted_f1")),
                    _fmt(row.get("test_roc_auc_macro")),
                    _fmt(row.get("val_macro_f1")),
                    _fmt(row.get("use_class_weights")),
                    _fmt(row.get("use_augmentation")),
                ]
            )
            + " |"
        )

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    args = parse_args()
    rows = sort_rows(load_rows(args.output_root), args.sort_metric)

    output_csv = _with_sort_metric_suffix(args.output_csv, args.sort_metric)
    output_md = _with_sort_metric_suffix(args.output_md, args.sort_metric)
    output_xlsx = _with_sort_metric_suffix(args.output_xlsx, args.sort_metric)

    write_csv(rows, output_csv)
    write_markdown(rows, output_md)
    write_xlsx(rows, output_xlsx)

    print(f"Runs compared : {len(rows)}")
    print(f"CSV written   : {output_csv}")
    print(f"Markdown      : {output_md}")
    print(f"Excel written : {output_xlsx}")
    if rows:
        best = rows[0]

        print(f"Best run      : {best['run_id']}")
        print(f"Best model    : {best['model_name']}")

        print(f"Sort metric   : {args.sort_metric}")
        print(
            f"{args.sort_metric}: "
            f"{_fmt(best.get(args.sort_metric))}"
        )

        print(f"Best accuracy : {_fmt(best.get('test_accuracy'))}")
        print(f"Best macro F1 : {_fmt(best.get('test_macro_f1'))}")
        print(f"Best recall   : {_fmt(best.get('test_macro_recall'))}")
        print(f"Best precision: {_fmt(best.get('test_macro_precision'))}")


if __name__ == "__main__":
    main()
